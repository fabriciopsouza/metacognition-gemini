import openpyxl

filepath = r'F:\Downloads\DGO-OPER-360\DGO-OPER-360_2026_v2.xlsx'
wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

for sheet_name in ['T4 Realizado', 'TMPU Realizado']:
    if sheet_name in wb.sheetnames:
        print(f"\n--- Aba: {sheet_name} ---")
        sheet = wb[sheet_name]
        for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True)):
            if any(row):
                print(f"Linha {idx+1}: {row[:15]}") # Print first 15 columns
    else:
        print(f"\nAba {sheet_name} não encontrada!")
wb.close()
