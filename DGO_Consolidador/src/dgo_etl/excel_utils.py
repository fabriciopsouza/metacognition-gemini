import openpyxl
from typing import List, Dict, Any, Optional

def iter_excel_sheet(filepath: str, sheet_name: str, header_keywords: List[str], max_header_row: int = 200) -> List[Dict[str, Any]]:
    """
    Busca dinamicamente a linha de cabeçalho varrendo até 'max_header_row' linhas.
    Mapeia os índices exatos das colunas para suportar deslocamentos verticais 
    e horizontais (tabelas flutuantes no Excel).
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Aba '{sheet_name}' não encontrada em {filepath}")
        
    sheet = wb[sheet_name]
    header_row_idx = -1
    headers_map = {}
    last_seen_col_values = {}
    
    # Busca cabeçalho
    for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_header_row, values_only=True), 1):
        if not any(row):
            continue
            
        # Rastreia o último valor preenchido de cada coluna (resolve merges verticais)
        for c_idx, c_val in enumerate(row):
            if c_val is not None and str(c_val).strip() != "":
                last_seen_col_values[c_idx] = str(c_val).strip()
                
        row_strs = [str(c).strip().lower() for c in row if c is not None]
        
        # Valida se todas as keywords exigidas existem exatamente como valor da célula
        if all(kw.lower() in row_strs for kw in header_keywords):
            header_row_idx = r_idx
            # Mapeia índice da coluna herdando o último valor preenchido acima (Opção A - HITL)
            headers_map = dict(last_seen_col_values)
            break
            
    if header_row_idx == -1:
        wb.close()
        raise ValueError(f"Cabeçalho com {header_keywords} não encontrado nas {max_header_row} primeiras linhas da aba {sheet_name}.")
        
    # Extrai os dados
    data = []
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not any(row):
            continue
            
        # Se a linha só tem nulos nas colunas de interesse, ignora
        if all(row[c_idx] is None for c_idx in headers_map.keys() if c_idx < len(row)):
            continue
            
        row_dict = {}
        for c_idx, header_name in headers_map.items():
            if c_idx < len(row):
                row_dict[header_name] = row[c_idx]
            else:
                row_dict[header_name] = None
                
        data.append(row_dict)
        
    wb.close()
    return data
